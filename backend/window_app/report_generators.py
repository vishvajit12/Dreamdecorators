"""
report_generators.py
────────────────────
Generates downloadable PDF and Excel reports for the Window & Door software.

Reports:
1. Quotation PDF
2. Complete Material BOQ PDF + Excel
3. Bar Optimisation (Cutting) Report PDF + Excel

Dependencies: reportlab, openpyxl
"""

from __future__ import annotations
import io
from datetime import date
from typing import List

# ─── PDF helpers (ReportLab) ──────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

# ─── Excel ────────────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from window_app.calculation_engine import (
    generate_boq, optimise_bars, build_quotation_summary,
    BOQLine, BarOptimisationResult
)

# ─── Brand colours ────────────────────────────────────────────────────────────
PRIMARY = colors.HexColor('#1a3c5e')       # deep navy
ACCENT = colors.HexColor('#e67e22')        # orange
LIGHT_BG = colors.HexColor('#f0f4f8')
MID_BG = colors.HexColor('#d5e3f0')
WHITE = colors.white
BLACK = colors.black
GREEN = colors.HexColor('#27ae60')
RED = colors.HexColor('#e74c3c')


def _header_table(project, report_title: str) -> Table:
    """Create a standard report header table with company info."""
    data = [
        [
            Paragraph(
                '<font size="16" color="#1a3c5e"><b>D Sign Design</b></font><br/>'
                '<font size="9" color="#555">Innovation to Production</font>',
                ParagraphStyle('h', fontName='Helvetica')
            ),
            Paragraph(
                f'<font size="14"><b>{report_title}</b></font><br/>'
                f'<font size="9">Project: {project.project_name}</font><br/>'
                f'<font size="9">Date: {date.today().strftime("%d/%m/%Y")}</font>',
                ParagraphStyle('h2', fontName='Helvetica', alignment=TA_RIGHT)
            ),
        ]
    ]
    t = Table(data, colWidths=[90 * mm, 100 * mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), LIGHT_BG),
        ('BOX', (0, 0), (-1, -1), 1.5, PRIMARY),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    return t


# ─── 1. Quotation PDF ─────────────────────────────────────────────────────────

def generate_quotation_pdf(project) -> bytes:
    """
    Generate a professional customer-facing quotation PDF.

    Returns:
        bytes: PDF file content
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=15 * mm, bottomMargin=15 * mm,
        leftMargin=15 * mm, rightMargin=15 * mm
    )
    styles = getSampleStyleSheet()
    normal = ParagraphStyle('n', fontName='Helvetica', fontSize=9)
    bold = ParagraphStyle('b', fontName='Helvetica-Bold', fontSize=9)
    heading = ParagraphStyle('hd', fontName='Helvetica-Bold', fontSize=11, textColor=PRIMARY)

    elements = []

    # Header
    elements.append(_header_table(project, 'QUOTATION'))
    elements.append(Spacer(1, 8 * mm))

    # Customer info
    cust_data = [
        ['Customer:', project.customer_name, 'Quote Date:', date.today().strftime('%d/%m/%Y')],
        ['Address:', project.customer_address or '-', 'Project:', project.project_name],
        ['Phone:', project.customer_phone or '-', 'Email:', project.customer_email or '-'],
    ]
    ct = Table(cust_data, colWidths=[25 * mm, 75 * mm, 30 * mm, 60 * mm])
    ct.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (-1, -1), LIGHT_BG),
        ('GRID', (0, 0), (-1, -1), 0.5, MID_BG),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(ct)
    elements.append(Spacer(1, 6 * mm))

    # Items table
    elements.append(Paragraph('ITEMS', heading))
    elements.append(Spacer(1, 2 * mm))

    item_header = ['#', 'Code', 'Description', 'Size (W×H mm)', 'Typology', 'Glass', 'Mesh', 'Qty']
    item_rows = [item_header]
    for i, item in enumerate(project.items.select_related('typology', 'glass_type', 'finish').all(), 1):
        item_rows.append([
            str(i),
            item.code,
            item.notes or '-',
            f"{item.width:.0f} × {item.height:.0f}",
            item.typology.display_name,
            item.glass_type.name,
            'Yes' if item.has_mesh else 'No',
            str(item.quantity),
        ])

    it = Table(item_rows, colWidths=[8 * mm, 15 * mm, 35 * mm, 28 * mm, 35 * mm, 28 * mm, 12 * mm, 10 * mm])
    it.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
        ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, LIGHT_BG]),
        ('GRID', (0, 0), (-1, -1), 0.5, MID_BG),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(it)
    elements.append(Spacer(1, 8 * mm))

    # Financials
    summary = build_quotation_summary(project)
    elements.append(Paragraph('COST SUMMARY', heading))
    elements.append(Spacer(1, 2 * mm))

    fin_data = [
        ['Profile / Aluminium Material', f'₹ {summary.subtotal_profiles:,.2f}'],
        ['Hardware & Accessories', f'₹ {summary.subtotal_hardware:,.2f}'],
        ['Glass Supply', f'₹ {summary.subtotal_glass:,.2f}'],
        ['Labour & Fabrication', f'₹ {summary.subtotal_labour:,.2f}'],
        ['Sub-Total', f'₹ {summary.subtotal_before_discount:,.2f}'],
        [f'Discount ({summary.discount_percent:.1f}%)', f'- ₹ {summary.discount_amount:,.2f}'],
        ['Amount after Discount', f'₹ {summary.subtotal_after_discount:,.2f}'],
        [f'GST ({summary.gst_percent:.0f}%)', f'₹ {summary.gst_amount:,.2f}'],
        ['GRAND TOTAL', f'₹ {summary.grand_total:,.2f}'],
    ]

    ft = Table(fin_data, colWidths=[120 * mm, 70 * mm])
    ft.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 4), (1, 4), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (1, -1), 11),
        ('BACKGROUND', (0, 4), (-1, 4), MID_BG),
        ('BACKGROUND', (0, -1), (-1, -1), PRIMARY),
        ('TEXTCOLOR', (0, -1), (-1, -1), WHITE),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, MID_BG),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(ft)

    elements.append(Spacer(1, 10 * mm))
    elements.append(Paragraph(
        '<font size="8" color="#888">This quotation is valid for 30 days. '
        'Prices are subject to change without prior notice. '
        'Terms & Conditions apply. | D Sign Design | www.dsigndesign.co.in</font>',
        ParagraphStyle('footer', fontName='Helvetica', fontSize=8, alignment=TA_CENTER)
    ))

    doc.build(elements)
    return buffer.getvalue()


# ─── 2. BOQ PDF ───────────────────────────────────────────────────────────────

def generate_boq_pdf(project) -> bytes:
    """Generate detailed material BOQ PDF with all cut lengths."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=15 * mm, bottomMargin=15 * mm,
        leftMargin=12 * mm, rightMargin=12 * mm
    )
    heading = ParagraphStyle('hd', fontName='Helvetica-Bold', fontSize=11, textColor=PRIMARY)
    elements = []

    elements.append(_header_table(project, 'MATERIAL BOQ – DETAILED CUTTING CHART'))
    elements.append(Spacer(1, 6 * mm))

    profile_lines, hardware_lines, glass_lines = generate_boq(project)

    def make_table(rows, col_widths, header_color=PRIMARY):
        t = Table(rows, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, LIGHT_BG]),
            ('GRID', (0, 0), (-1, -1), 0.5, MID_BG),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        return t

    # ── Profiles ─────────────────────────────────────────────────────────────
    elements.append(Paragraph('PROFILE CUTTING SCHEDULE', heading))
    elements.append(Spacer(1, 2 * mm))

    ph = ['Item', 'Profile', 'Direction', 'Cut Length (mm)', 'Qty', 'Total Length (mm)', 'Rate/m (₹)', 'Amount (₹)']
    prows = [ph]
    for l in profile_lines:
        prows.append([
            l.item_code, l.description.split('(')[0].strip(), l.description.split('(')[-1].replace(')', ''),
            f'{l.cut_length_mm:.1f}', str(l.quantity),
            f'{l.total_length_mm:.1f}', f'{l.unit_price:.2f}', f'{l.total_price:.2f}'
        ])
    prows.append([
        '', '', '', 'TOTAL', '', f'{sum(l.total_length_mm for l in profile_lines):.1f}',
        '', f'{sum(l.total_price for l in profile_lines):.2f}'
    ])
    elements.append(make_table(prows, [15 * mm, 42 * mm, 22 * mm, 28 * mm, 10 * mm, 30 * mm, 20 * mm, 22 * mm]))
    elements.append(Spacer(1, 6 * mm))

    # ── Hardware ─────────────────────────────────────────────────────────────
    elements.append(Paragraph('HARDWARE & ACCESSORIES BOQ', heading))
    elements.append(Spacer(1, 2 * mm))

    hh = ['Item', 'Description', 'Unit', 'Qty', 'Unit Price (₹)', 'Amount (₹)']
    hrows = [hh]
    for l in hardware_lines:
        hrows.append([l.item_code, l.description, 'nos', str(l.quantity), f'{l.unit_price:.2f}', f'{l.total_price:.2f}'])
    hrows.append(['', 'TOTAL', '', '', '', f'{sum(l.total_price for l in hardware_lines):.2f}'])
    elements.append(make_table(hrows, [15 * mm, 65 * mm, 15 * mm, 20 * mm, 30 * mm, 30 * mm]))
    elements.append(Spacer(1, 6 * mm))

    # ── Glass ─────────────────────────────────────────────────────────────────
    elements.append(Paragraph('GLASS SCHEDULE', heading))
    elements.append(Spacer(1, 2 * mm))

    gh = ['Item', 'Glass / Finish', 'Qty', 'Area (sq.ft)', 'Rate/sq.ft (₹)', 'Amount (₹)']
    grows = [gh]
    for l in glass_lines:
        grows.append([
            l.item_code, l.description, str(l.quantity),
            f'{l.total_length_mm:.3f}', f'{l.unit_price:.2f}', f'{l.total_price:.2f}'
        ])
    grows.append(['', 'TOTAL', '', f'{sum(l.total_length_mm for l in glass_lines):.3f}', '',
                  f'{sum(l.total_price for l in glass_lines):.2f}'])
    elements.append(make_table(grows, [15 * mm, 65 * mm, 15 * mm, 25 * mm, 30 * mm, 25 * mm]))

    doc.build(elements)
    return buffer.getvalue()


# ─── 3. Bar Optimisation PDF ──────────────────────────────────────────────────

def generate_bar_optimisation_pdf(project) -> bytes:
    """Generate bar cutting optimisation report PDF."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=15 * mm, bottomMargin=15 * mm,
        leftMargin=12 * mm, rightMargin=12 * mm
    )
    heading = ParagraphStyle('hd', fontName='Helvetica-Bold', fontSize=11, textColor=PRIMARY)
    sub = ParagraphStyle('sub', fontName='Helvetica-Bold', fontSize=9, textColor=ACCENT)
    elements = []

    elements.append(_header_table(project, 'BAR OPTIMISATION – CUTTING REPORT'))
    elements.append(Spacer(1, 6 * mm))

    results = optimise_bars(project)

    for result in results:
        elements.append(Paragraph(f'Profile: {result.profile_name}', heading))

        # Summary row
        summary_data = [
            ['Bar Length', 'Total Bars', 'Total Cut (mm)', 'Total Waste (mm)', 'Efficiency'],
            [
                f'{result.bar_length:.0f} mm',
                str(result.total_bars),
                f'{result.total_cut_length_mm:.1f}',
                f'{result.total_waste_mm:.1f}',
                f'{result.efficiency_percent:.1f}%',
            ]
        ]
        st = Table(summary_data, colWidths=[35 * mm, 25 * mm, 35 * mm, 35 * mm, 25 * mm])
        eff_color = GREEN if result.efficiency_percent >= 75 else ACCENT if result.efficiency_percent >= 60 else RED
        st.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
            ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (4, 1), (4, 1), eff_color),
            ('TEXTCOLOR', (4, 1), (4, 1), WHITE),
            ('GRID', (0, 0), (-1, -1), 0.5, MID_BG),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(st)
        elements.append(Spacer(1, 3 * mm))

        # Per-bar detail
        for bar in result.bars:
            elements.append(Paragraph(
                f'Bar #{bar.bar_number}  |  Used: {bar.used_mm:.1f}mm  |  Waste: {bar.waste_mm:.1f}mm',
                sub
            ))
            cut_header = ['Cut #', 'Label', 'Length (mm)', 'Cumulative (mm)']
            cut_rows = [cut_header]
            cumulative = 0.0
            for j, (label, length) in enumerate(bar.cuts, 1):
                cumulative += length + bar.kerf_mm
                cut_rows.append([str(j), label, f'{length:.1f}', f'{cumulative:.1f}'])
            ct = Table(cut_rows, colWidths=[15 * mm, 50 * mm, 35 * mm, 35 * mm])
            ct.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), MID_BG),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, LIGHT_BG]),
                ('GRID', (0, 0), (-1, -1), 0.5, MID_BG),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            elements.append(ct)
            elements.append(Spacer(1, 2 * mm))

        elements.append(Spacer(1, 6 * mm))

    doc.build(elements)
    return buffer.getvalue()


# ─── Excel BOQ ────────────────────────────────────────────────────────────────

def generate_boq_excel(project) -> bytes:
    """Generate Material BOQ as Excel workbook with multiple sheets."""
    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='1a3c5e')
    sub_fill = PatternFill('solid', fgColor='d5e3f0')
    total_fill = PatternFill('solid', fgColor='e67e22')
    total_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    normal_font = Font(name='Calibri', size=9)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center')
    thin = Side(style='thin', color='b0c4de')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws, row_num, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

    def style_data_row(ws, row_num, max_col, alt=False):
        fill = PatternFill('solid', fgColor='f0f4f8') if alt else None
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = normal_font
            cell.alignment = center
            cell.border = border
            if fill:
                cell.fill = fill

    profile_lines, hardware_lines, glass_lines = generate_boq(project)

    # ── Sheet 1: Profiles ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Profile Cutting'

    ws1.merge_cells('A1:H1')
    ws1['A1'] = f'PROFILE CUTTING SCHEDULE | {project.project_name}'
    ws1['A1'].font = Font(name='Calibri', bold=True, size=12, color='1a3c5e')
    ws1['A1'].alignment = center

    headers = ['Item Code', 'Profile Name', 'Direction', 'Cut Length (mm)', 'Qty', 'Total Length (mm)', 'Rate/m (₹)', 'Amount (₹)']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=2, column=col, value=h)
    style_header_row(ws1, 2, len(headers))

    for i, l in enumerate(profile_lines, 3):
        vals = [l.item_code, l.description.split('(')[0].strip(),
                l.description.split('(')[-1].replace(')', ''),
                round(l.cut_length_mm, 1), l.quantity,
                round(l.total_length_mm, 1), l.unit_price, round(l.total_price, 2)]
        for col, v in enumerate(vals, 1):
            ws1.cell(row=i, column=col, value=v)
        style_data_row(ws1, i, len(headers), alt=(i % 2 == 0))

    total_row = len(profile_lines) + 3
    ws1.cell(row=total_row, column=4, value='TOTAL')
    ws1.cell(row=total_row, column=6, value=round(sum(l.total_length_mm for l in profile_lines), 1))
    ws1.cell(row=total_row, column=8, value=round(sum(l.total_price for l in profile_lines), 2))
    for col in range(1, len(headers) + 1):
        cell = ws1.cell(row=total_row, column=col)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.alignment = center

    col_widths1 = [12, 35, 18, 18, 8, 22, 16, 16]
    for col, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 2: Hardware ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Hardware BOQ')
    ws2.merge_cells('A1:F1')
    ws2['A1'] = f'HARDWARE & ACCESSORIES BOQ | {project.project_name}'
    ws2['A1'].font = Font(name='Calibri', bold=True, size=12, color='1a3c5e')
    ws2['A1'].alignment = center

    hw_headers = ['Item Code', 'Description', 'Unit', 'Qty', 'Unit Price (₹)', 'Amount (₹)']
    for col, h in enumerate(hw_headers, 1):
        ws2.cell(row=2, column=col, value=h)
    style_header_row(ws2, 2, len(hw_headers))

    for i, l in enumerate(hardware_lines, 3):
        vals = [l.item_code, l.description, 'nos', l.quantity, l.unit_price, round(l.total_price, 2)]
        for col, v in enumerate(vals, 1):
            ws2.cell(row=i, column=col, value=v)
        style_data_row(ws2, i, len(hw_headers), alt=(i % 2 == 0))

    hw_total = len(hardware_lines) + 3
    ws2.cell(row=hw_total, column=2, value='TOTAL')
    ws2.cell(row=hw_total, column=6, value=round(sum(l.total_price for l in hardware_lines), 2))
    for col in range(1, len(hw_headers) + 1):
        cell = ws2.cell(row=hw_total, column=col)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.alignment = center

    for col, w in enumerate([12, 45, 10, 10, 18, 18], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 3: Glass ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Glass Schedule')
    ws3.merge_cells('A1:F1')
    ws3['A1'] = f'GLASS SCHEDULE | {project.project_name}'
    ws3['A1'].font = Font(name='Calibri', bold=True, size=12, color='1a3c5e')
    ws3['A1'].alignment = center

    gl_headers = ['Item Code', 'Glass / Finish', 'Qty', 'Area (sq.ft)', 'Rate/sqft (₹)', 'Amount (₹)']
    for col, h in enumerate(gl_headers, 1):
        ws3.cell(row=2, column=col, value=h)
    style_header_row(ws3, 2, len(gl_headers))

    for i, l in enumerate(glass_lines, 3):
        vals = [l.item_code, l.description, l.quantity,
                round(l.total_length_mm, 3), l.unit_price, round(l.total_price, 2)]
        for col, v in enumerate(vals, 1):
            ws3.cell(row=i, column=col, value=v)
        style_data_row(ws3, i, len(gl_headers), alt=(i % 2 == 0))

    gl_total = len(glass_lines) + 3
    ws3.cell(row=gl_total, column=2, value='TOTAL')
    ws3.cell(row=gl_total, column=4, value=round(sum(l.total_length_mm for l in glass_lines), 3))
    ws3.cell(row=gl_total, column=6, value=round(sum(l.total_price for l in glass_lines), 2))
    for col in range(1, len(gl_headers) + 1):
        cell = ws3.cell(row=gl_total, column=col)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.alignment = center

    for col, w in enumerate([12, 45, 8, 16, 18, 18], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 4: Summary ──────────────────────────────────────────────────────
    ws4 = wb.create_sheet('Quotation Summary')
    summary = build_quotation_summary(project)

    ws4['A1'] = 'QUOTATION SUMMARY'
    ws4['A1'].font = Font(name='Calibri', bold=True, size=14, color='1a3c5e')

    rows = [
        ('', ''),
        ('Project', project.project_name),
        ('Customer', project.customer_name),
        ('Date', str(date.today())),
        ('', ''),
        ('Profile Material', f'₹ {summary.subtotal_profiles:,.2f}'),
        ('Hardware', f'₹ {summary.subtotal_hardware:,.2f}'),
        ('Glass', f'₹ {summary.subtotal_glass:,.2f}'),
        ('Labour', f'₹ {summary.subtotal_labour:,.2f}'),
        ('Sub-Total', f'₹ {summary.subtotal_before_discount:,.2f}'),
        (f'Discount ({summary.discount_percent:.1f}%)', f'-₹ {summary.discount_amount:,.2f}'),
        ('After Discount', f'₹ {summary.subtotal_after_discount:,.2f}'),
        (f'GST ({summary.gst_percent:.0f}%)', f'₹ {summary.gst_amount:,.2f}'),
        ('GRAND TOTAL', f'₹ {summary.grand_total:,.2f}'),
    ]

    for i, (label, value) in enumerate(rows, 2):
        ws4.cell(row=i, column=1, value=label).font = Font(name='Calibri', bold=True, size=10)
        ws4.cell(row=i, column=2, value=value).font = Font(name='Calibri', size=10)

    # Grand total highlight
    grand_row = len(rows) + 1
    ws4.cell(row=grand_row, column=1).fill = PatternFill('solid', fgColor='1a3c5e')
    ws4.cell(row=grand_row, column=1).font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    ws4.cell(row=grand_row, column=2).fill = PatternFill('solid', fgColor='1a3c5e')
    ws4.cell(row=grand_row, column=2).font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')

    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 25

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ─── Excel Bar Optimisation ───────────────────────────────────────────────────

def generate_bar_optimisation_excel(project) -> bytes:
    """Generate bar optimisation report as Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Bar Optimisation'

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='1a3c5e')
    sub_fill = PatternFill('solid', fgColor='d5e3f0')
    thin = Side(style='thin', color='b0c4de')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    results = optimise_bars(project)

    row = 1
    ws.cell(row=row, column=1, value=f'BAR OPTIMISATION REPORT | {project.project_name}').font = \
        Font(name='Calibri', bold=True, size=14, color='1a3c5e')
    row += 2

    for result in results:
        # Profile header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=f'Profile: {result.profile_name}').font = \
            Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = header_fill
            ws.cell(row=row, column=col).border = border
        row += 1

        # Summary
        for col, (label, val) in enumerate([
            ('Bar Length (mm)', result.bar_length),
            ('Total Bars', result.total_bars),
            ('Total Cut (mm)', result.total_cut_length_mm),
            ('Waste (mm)', result.total_waste_mm),
            ('Efficiency %', result.efficiency_percent),
        ], 1):
            ws.cell(row=row, column=col, value=label).fill = sub_fill
            ws.cell(row=row, column=col).font = Font(name='Calibri', bold=True, size=9)
            ws.cell(row=row + 1, column=col, value=val).font = Font(name='Calibri', size=9)
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row + 1, column=col).border = border
            ws.cell(row=row, column=col).alignment = center
            ws.cell(row=row + 1, column=col).alignment = center
        row += 2

        # Per-bar cuts
        for bar in result.bars:
            ws.cell(row=row, column=1, value=f'Bar #{bar.bar_number}').font = \
                Font(name='Calibri', bold=True, size=9)
            ws.cell(row=row, column=2, value=f'Used: {bar.used_mm:.1f}mm').font = Font(name='Calibri', size=9)
            ws.cell(row=row, column=3, value=f'Waste: {bar.waste_mm:.1f}mm').font = Font(name='Calibri', size=9)
            row += 1

            for j, (label, length) in enumerate(bar.cuts, 1):
                ws.cell(row=row, column=2, value=f'Cut {j}')
                ws.cell(row=row, column=3, value=label)
                ws.cell(row=row, column=4, value=round(length, 1))
                ws.cell(row=row, column=5, value='mm')
                for col in range(2, 6):
                    ws.cell(row=row, column=col).font = Font(name='Calibri', size=9)
                    ws.cell(row=row, column=col).border = border
                row += 1
            row += 1

        row += 2

    for col, w in enumerate([20, 15, 25, 20, 15, 15], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
